# -*- coding: utf-8 -*-
"""
產出兩張 A4 橫式服務簽收單，給洗毛這件小事店內手寫使用。

執行：
    python generate_forms.py

產出（與此 .py 同資料夾）：
    - 散客服務簽收單.xlsx       → 對應檔案 A「服務紀錄」分頁
    - 儲值客戶服務簽收單.xlsx   → 對應檔案 B「交易明細」分頁

設計原則：
    - 欄位順序 100% 對齊目標 Sheets（Cowork 識別後從左到右就是 row 順序）
    - 能勾的不要寫（服務項目 / 付款方式 預填 ☐ checkbox，員工只勾 ✓）
    - 日期、美容師放頁首一次寫，每列不重寫 → 少一堆手寫錯誤
    - 簽收欄寬給足，A4 橫式直接列印不用調

修改提示（改完重跑就好）：
    - 改店名      → SHOP_NAME
    - 改服務項目  → SERVICE_OPTIONS
    - 改付款方式  → PAYMENT_OPTIONS
    - 改列數      → ROW_COUNT
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ===== 可調整 =====
SHOP_NAME = "洗毛這件小事"
SERVICE_OPTIONS = ["洗澡", "全美容", "剪指甲", "SPA"]
PAYMENT_OPTIONS = ["現金", "匯款", "LINE Pay"]

# 主版本列數 → 主檔名 (散客服務簽收單.xlsx)
# 額外產的列數 → 加後綴檔名 (散客服務簽收單_20列.xlsx)
PRIMARY_ROW_COUNT = 15
EXTRA_ROW_COUNTS = []     # 想多產備援列數版本就加進來，例如 [20]；每組產 2 個檔（檔名加後綴 _NN列）
# ==================

# A4 橫式可用主體高度（扣掉標題/頁首/表頭/頁尾後給列用的 pt）
BODY_AVAILABLE_PT = 417

FONT_NAME = "微軟正黑體"
TITLE_FONT = Font(name=FONT_NAME, size=18, bold=True)
SUBHEAD_FONT = Font(name=FONT_NAME, size=11)
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME, size=11)
CHECKBOX_FONT = Font(name=FONT_NAME, size=10)
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="666666")

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
ZEBRA_FILL = PatternFill("solid", fgColor="F2F7FB")

THIN = Side(style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)


def build_service_options_text(compact=False):
    if compact:
        # 單行緊湊：☐洗澡 ☐全美容 ☐剪指甲 ☐SPA ☐其他___
        return "  ".join(f"☐{x}" for x in SERVICE_OPTIONS) + "  ☐其他___"
    # 兩兩一組換行 + 「其他___」
    lines = []
    for i in range(0, len(SERVICE_OPTIONS), 2):
        pair = SERVICE_OPTIONS[i:i + 2]
        lines.append("    ".join(f"☐ {x}" for x in pair))
    lines.append("☐ 其他：__________")
    return "\n".join(lines)


def build_payment_options_text(compact=False):
    if compact:
        return " ".join(f"☐{x}" for x in PAYMENT_OPTIONS)
    return "    ".join(f"☐ {x}" for x in PAYMENT_OPTIONS)


def calc_row_height(row_count):
    """依列數動態算主體列高（pt），保證一頁印得下。"""
    return max(20, BODY_AVAILABLE_PT / row_count)


def use_compact_layout(row_count):
    """列高小於 30 pt 改用緊湊（單行）checkbox。"""
    return calc_row_height(row_count) < 30


def setup_sheet(ws: Worksheet, title: str, columns, sub_label: str,
                row_count: int, prefill=None):
    """
    columns: list of (key, header_label, width_chars)
    row_count: 主體列數（會依此自動算行高 / 切換 checkbox 緊湊模式）
    prefill: dict {key: text_to_prefill_in_each_body_row}
    """
    prefill = prefill or {}
    body_row_height = calc_row_height(row_count)
    keys = [c[0] for c in columns]
    headers = [c[1] for c in columns]
    widths = [c[2] for c in columns]
    n = len(columns)

    # 第 1 列：標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    title_cell = ws.cell(row=1, column=1, value=f"🐾 {SHOP_NAME} — {title}")
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 34

    # 第 2 列：日期 + 美容師（半半切）
    half = max(1, n // 2)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=half)
    c = ws.cell(row=2, column=1, value="日期：______ 年 ____ 月 ____ 日")
    c.font = SUBHEAD_FONT
    c.alignment = LEFT
    ws.merge_cells(start_row=2, start_column=half + 1, end_row=2, end_column=n)
    c = ws.cell(row=2, column=half + 1,
                value=f"美容師：________________      （此單用途：{sub_label}）")
    c.font = SUBHEAD_FONT
    c.alignment = LEFT
    ws.row_dimensions[2].height = 26

    # 第 3 列：表頭
    for ci, header in enumerate(headers, start=1):
        c = ws.cell(row=3, column=ci, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]
    ws.row_dimensions[3].height = 38

    # 第 4 列起：主體
    for ri in range(row_count):
        excel_row = 4 + ri
        ws.row_dimensions[excel_row].height = body_row_height
        for ci in range(1, n + 1):
            key = keys[ci - 1]
            if key == "no":
                value = ri + 1
                cell = ws.cell(row=excel_row, column=ci, value=value)
                cell.alignment = CENTER
                cell.font = BODY_FONT
            elif key in prefill:
                cell = ws.cell(row=excel_row, column=ci, value=prefill[key])
                cell.alignment = LEFT_TOP
                cell.font = CHECKBOX_FONT
            else:
                cell = ws.cell(row=excel_row, column=ci, value="")
                cell.alignment = LEFT
                cell.font = BODY_FONT
            cell.border = BORDER
            if ri % 2 == 1:
                cell.fill = ZEBRA_FILL

    # 頁尾備註
    note_row = 4 + row_count
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=n)
    c = ws.cell(
        row=note_row, column=1,
        value="※ 收班拍照傳 Cowork KEY IN ｜ 電話寫滿 10 碼開頭 0 ｜ 金額只寫純數字（例 800，不寫 NT$ / 不寫「元」） ｜ 勾選用 ✓ 或塗滿，避免畫斜線",
    )
    c.font = NOTE_FONT
    c.alignment = LEFT
    ws.row_dimensions[note_row].height = 22

    # 列印設定：A4 橫式 fit-to-page
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.print_options.horizontalCentered = True
    ws.print_area = f"A1:{get_column_letter(n)}{note_row}"


def make_walk_in_form(row_count: int, file_suffix: str = ""):
    """散客單（對應檔案 A「服務紀錄」分頁）

    Sheets 欄位順序：
        A 日期 | B 主人電話 | C 寵物名 | D 服務項目 | E 金額 | F 付款方式 | G 美容師 | H 備註
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "散客服務簽收單"

    compact = use_compact_layout(row_count)
    # 緊湊模式時服務項目 / 付款方式 cell 縮窄、其他欄稍微讓位
    columns = [
        ("no", "#", 4),
        ("phone", "主人電話\n(10 碼開頭 0)", 17),
        ("pet", "寵物名", 11),
        ("service", "服務項目（勾選）", 28 if compact else 24),
        ("amount", "金額\n(NT$)", 10),
        ("payment", "付款方式（勾選）" if compact else "付款方式\n（勾選）", 18 if compact else 17),
        ("sign", "簽收", 14),
        ("note", "備註", 14 if compact else 16),
    ]
    prefill = {
        "service": build_service_options_text(compact=compact),
        "payment": build_payment_options_text(compact=compact),
    }
    setup_sheet(ws, "散客服務簽收單", columns,
                sub_label="未預約散客 / 一般非儲值客",
                row_count=row_count, prefill=prefill)

    out = os.path.join(OUT_DIR, f"散客服務簽收單{file_suffix}.xlsx")
    wb.save(out)
    print(f"OK 寫好：{out}  (列數={row_count}, 列高={calc_row_height(row_count):.1f}pt, "
          f"checkbox={'單行緊湊' if compact else '雙行'})")


def make_credit_form(row_count: int, file_suffix: str = ""):
    """儲值客戶單（對應檔案 B「交易明細」分頁）

    Sheets 欄位順序：
        A 電話 | B 客戶姓名 | C 日期 | D 儲值金額 | E 消費項目 | F 消費金額 | G 餘額 | H 簽名 | I 備註
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "儲值客戶服務簽收單"

    compact = use_compact_layout(row_count)
    columns = [
        ("no", "#", 4),
        ("phone", "主人電話\n(10 碼開頭 0)", 17),
        ("name", "客戶姓名", 11),
        ("service", "服務項目（勾選）", 26 if compact else 22),
        ("amount", "消費金額", 10),
        ("topup", "儲值金額\n(新增儲值才填)", 12),
        ("balance", "扣後餘額", 10),
        ("sign", "簽收", 13),
        ("note", "備註", 12 if compact else 13),
    ]
    prefill = {
        "service": build_service_options_text(compact=compact),
    }
    setup_sheet(ws, "儲值客戶服務簽收單", columns,
                sub_label="已儲值客 / 餘額扣抵服務",
                row_count=row_count, prefill=prefill)

    out = os.path.join(OUT_DIR, f"儲值客戶服務簽收單{file_suffix}.xlsx")
    wb.save(out)
    print(f"OK 寫好：{out}  (列數={row_count}, 列高={calc_row_height(row_count):.1f}pt, "
          f"checkbox={'單行緊湊' if compact else '雙行'})")


if __name__ == "__main__":
    print(f"店名         : {SHOP_NAME}")
    print(f"服務項目選項 : {', '.join(SERVICE_OPTIONS)}（其他可手寫）")
    print(f"付款方式選項 : {', '.join(PAYMENT_OPTIONS)}")
    print(f"主版本列數   : {PRIMARY_ROW_COUNT}")
    print(f"額外列數版本 : {EXTRA_ROW_COUNTS}")
    print()

    # 主版本（無後綴）
    make_walk_in_form(PRIMARY_ROW_COUNT, "")
    make_credit_form(PRIMARY_ROW_COUNT, "")

    # 額外列數（檔名加後綴 _NN列）
    for rc in EXTRA_ROW_COUNTS:
        suffix = f"_{rc}列"
        make_walk_in_form(rc, suffix)
        make_credit_form(rc, suffix)

    print()
    print("列印建議：A4 橫式、彩色、邊界已設好、不用縮放")
    print("改格子請編輯本檔上方常數區後重跑")
