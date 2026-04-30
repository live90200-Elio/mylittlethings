# -*- coding: utf-8 -*-
"""
把「客戶儲值單_完整版.xlsx」轉成 Google Sheets「寵物店儲值帳本」可以直接附加的 CSV。

跟 export_initial_8_to_csv.py 的差別：
  - export_initial_8_to_csv.py 從 create_储值单.py 讀現有 8 位客戶
  - 這支從 xlsx 直接讀，未來 5/15、5/30 再來新一批時也用這支

執行方式（預設讀 4/30 那份，可改檔名）：
    python export_xlsx_to_csv.py "C:/Users/ZING/我的雲端硬碟/洗毛這件小事/客戶儲值單/430客戶儲值單_完整版.xlsx" 9

第 2 個參數 = 客戶總覽 # 起始流水號（既有 8 位 → 從 9 接續）

會產出（不蓋掉舊 CSV，加 _xlsx 後綴）：
    交易明細_xlsx.csv
    客戶總覽_xlsx.csv

匯入步驟（每個 CSV 各做一次）：
    1. 開「寵物店儲值帳本」對應分頁
    2. 檔案 → 匯入 → 上傳 → 選 CSV
    3. 匯入位置選「附加到目前的工作表」
    4. 分隔符號選「自動偵測」
    5. 按「匯入資料」
"""
import csv
import datetime
import io
import os
import re
import sys
from openpyxl import load_workbook

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# === 參數 ===
DEFAULT_XLSX = r"C:\Users\ZING\我的雲端硬碟\洗毛這件小事\客戶儲值單\430客戶儲值單_完整版.xlsx"
xlsx_path = sys.argv[1] if len(sys.argv) >= 2 else DEFAULT_XLSX
start_idx = int(sys.argv[2]) if len(sys.argv) >= 3 else 9

if not os.path.exists(xlsx_path):
    print(f"❌ 找不到 xlsx：{xlsx_path}")
    sys.exit(1)

OUT_DIR = os.path.dirname(os.path.abspath(__file__))


def fmt_value(v):
    """把儲存格值轉成乾淨字串。
    - None → ""
    - datetime → "m/d"（Excel 把 "4/25" 字串自動解成 datetime，要還原成 m/d）
    - float 整數值（例如 4700.0）→ 帶千分位 "4,700"
    - 其他都 strip 後直接 str()
    """
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return f"{v.month}/{v.day}"
    if isinstance(v, datetime.date):
        return f"{v.month}/{v.day}"
    if isinstance(v, float):
        if v.is_integer():
            return f"{int(v):,}"
        return str(v)
    if isinstance(v, int):
        return f"{v:,}"
    return str(v).strip()


def parse_header_row(info_row):
    """從第二列（'姓名：XXX' / None / '電話：XXX' / None / '寵物：XXX'）抽姓名/電話/寵物"""
    name = phone = pet = ""
    for cell in info_row:
        if not cell:
            continue
        s = str(cell)
        if s.startswith("姓名："):
            name = s.replace("姓名：", "").strip()
        elif s.startswith("電話："):
            phone = s.replace("電話：", "").strip()
        elif s.startswith("寵物："):
            pet = s.replace("寵物：", "").strip()
    # 「（未填）」「未填」一律存空字串，方便老闆事後 fillin
    if name in ("（未填）", "未填"):
        name = ""
    return name, phone, pet


def is_footer_row(row):
    """判斷是不是 '★ 綠色底=...' 那種說明列"""
    if not row:
        return True
    first = row[0]
    if first is None:
        return all(c is None for c in row)
    s = str(first).strip()
    return s.startswith("★") or s == ""


# === 讀 xlsx ===
wb = load_workbook(xlsx_path, data_only=True)
print(f"✅ 讀到 xlsx：{xlsx_path}")

customers = []
for sheet_name in wb.sheetnames:
    if sheet_name == "客戶總覽":
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        print(f"  ⚠️ 跳過 {sheet_name}：列數不足")
        continue
    name, phone, pet = parse_header_row(rows[1])
    if not phone:
        # fallback：sheet 名稱本身就是電話
        phone = sheet_name.replace(" - ", " / ")  # 雙電話 sheet 用 "-" 分隔，改 "/" 跟既有格式一致
    # 從 row index 3 (第 4 列) 開始讀資料；遇 footer 停
    tx_rows = []
    for r in rows[3:]:
        if is_footer_row(r):
            break
        # row 結構: 日期 | 儲值金額 | 消費項目 | 消費金額 | 餘額 | 簽名 | 備註
        tx_rows.append(tuple(fmt_value(c) for c in r[:7]))
        # padding（萬一原 sheet 不到 7 欄）
        while len(tx_rows[-1]) < 7:
            tx_rows[-1] = tx_rows[-1] + ("",)
    customers.append(
        {
            "sheet": sheet_name,
            "name": name,
            "phone": phone,
            "pet": pet,
            "rows": tx_rows,
        }
    )
    print(f"  ✓ {sheet_name}：{name or '（未填）'} / 寵物={pet} / {len(tx_rows)} 筆交易")

print(f"\n✅ 共處理 {len(customers)} 位客戶")

# === 寫 交易明細_xlsx.csv ===
ledger_path = os.path.join(OUT_DIR, "交易明細_xlsx.csv")
with open(ledger_path, "w", encoding="utf-8-sig", newline="") as f:
    w = csv.writer(f)
    for cust in customers:
        for r in cust["rows"]:
            date, topup, item, spent, balance, sign, note = r
            w.writerow(
                [
                    cust["phone"],  # A 電話
                    cust["name"],   # B 客戶姓名
                    date,           # C 日期
                    topup,          # D 儲值金額
                    item,           # E 消費項目
                    spent,          # F 消費金額
                    balance,        # G 餘額
                    sign,           # H 簽名
                    note,           # I 備註
                ]
            )
print(f"✅ 寫好：{ledger_path}")

# === 寫 客戶總覽_xlsx.csv ===
summary_path = os.path.join(OUT_DIR, "客戶總覽_xlsx.csv")
with open(summary_path, "w", encoding="utf-8-sig", newline="") as f:
    w = csv.writer(f)
    for offset, cust in enumerate(customers):
        # 找最後一筆有「餘額」的當目前餘額
        last_balance = ""
        for r in reversed(cust["rows"]):
            if r[4] and str(r[4]).strip():
                last_balance = r[4]
                break
        # 消費筆數 = 有消費金額的列
        spent_count = sum(1 for r in cust["rows"] if r[3] and str(r[3]).strip())
        # 最後消費日 = 最後有消費金額的日期
        last_spent_date = ""
        for r in reversed(cust["rows"]):
            if r[3] and str(r[3]).strip():
                last_spent_date = r[0]
                break
        w.writerow(
            [
                start_idx + offset,  # A #
                cust["name"],        # B 客戶姓名
                cust["phone"],       # C 電話
                cust["pet"],         # D 寵物
                last_balance,        # E 目前餘額
                spent_count,         # F 消費筆數
                last_spent_date,     # G 最後消費日
                "",                  # H 備註
            ]
        )
print(f"✅ 寫好：{summary_path}")

# === 統計 ===
total_tx = sum(len(c["rows"]) for c in customers)
print()
print(f"📊 總交易筆數：{total_tx}")
print(f"📋 總客戶數：{len(customers)}")
print()
print("📥 接下來：")
print("  1. 開『寵物店儲值帳本』Sheets")
print("  2. 切到「交易明細」分頁 → 檔案→匯入→上傳→選 交易明細_xlsx.csv → 附加到目前的工作表")
print("  3. 切到「客戶總覽」分頁 → 同上，選 客戶總覽_xlsx.csv")
print("  4. 完成後可在 Apps Script 跑 recomputeSummary 驗證餘額對得起來（注意它會清空寵物欄）")
